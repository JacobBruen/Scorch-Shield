import mysql.connector
import pandas as pd
from openpyxl import load_workbook

# MySQL connection - update credentials as needed
conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password="Your_Password",  # Replace with your MySQL root password
    database="DisasterReliefDB"
)
cursor = conn.cursor()

# Run update queries first
updates = [
    "SET SQL_SAFE_UPDATES = 0;",
    "UPDATE Resource SET Request_ID='REQ001' WHERE Resource_ID='R001';",
    "UPDATE Resource_Request SET Status = 'Completed' WHERE Status = 'In Progress';",
    "SET SQL_SAFE_UPDATES = 1;"
]
for upd in updates:
    cursor.execute(upd)
conn.commit()

# All SELECT queries
queries = {
    "Victims starting with M": """
        SELECT
          v.Victim_ID,
          v.Victim_Name,
          v.Age,
          v.Location,
          v.Shelter_ID,
          ar.Request_ID AS Aid_Request_ID,
          ar.Status AS Aid_Request_Status,
          rr.Resource_ID,
          rr.Resource_Name,
          rr.Amount AS Quantity_Requested,
          rr.Status AS Resource_Request_Status
        FROM Victim AS v
        LEFT JOIN Aid_Request AS ar ON v.Victim_ID = ar.Victim_ID
        LEFT JOIN Resource_Request AS rr ON ar.Request_ID = rr.Request_ID
            AND rr.Status IN ('In Progress','Completed')
        WHERE v.Victim_Name LIKE 'M%'
        LIMIT 1000;
    """,
    "Top 5 Shelters by Occupancy": """
        SELECT
          s.Shelter_ID,
          s.Shelter_Name,
          COUNT(v.Victim_ID) AS Victim_Count,
          s.Capacity,
          ROUND(COUNT(v.Victim_ID)/s.Capacity * 100, 2) AS Occupancy_Pct
        FROM Shelter s
        LEFT JOIN Victim v ON s.Shelter_ID = v.Shelter_ID
        GROUP BY s.Shelter_ID, s.Shelter_Name, s.Capacity
        ORDER BY Victim_Count DESC
        LIMIT 5;
    """,
    "Resource Completion Percentages": """
        SELECT
          r.Resource_ID,
          r.Resource_Name,
          d.Amount AS Donated_Amount,
          SUM(rr.Amount) AS Completed_Amount,
          ROUND(SUM(rr.Amount) / d.Amount * 100, 2) AS Percent_Completed
        FROM Resource_Request rr
        JOIN Resource r USING(Resource_ID)
        JOIN Donation d USING(Donation_ID)
        WHERE rr.Status = 'Completed'
        GROUP BY r.Resource_ID, r.Resource_Name, d.Amount
        ORDER BY r.Resource_ID;
    """,
    "Total Resources Requested": """
        SELECT
          rr.Resource_ID,
          r.Resource_Name,
          SUM(rr.Amount) AS Total_Requested
        FROM Resource_Request rr
        JOIN Resource r ON rr.Resource_ID = r.Resource_ID
        GROUP BY rr.Resource_ID, r.Resource_Name
        ORDER BY rr.Resource_ID;
    """,
    "Resources Under 50%": """
        SELECT
          r.Resource_ID,
          r.Resource_Name,
          r.Amount AS Current_Amount,
          d.Amount AS Total_Amount,
          ROUND(r.Amount / d.Amount * 100, 2) AS Percent_Remaining
        FROM Resource r
        JOIN Donation d ON r.Donation_ID = d.Donation_ID
        WHERE r.Amount < 0.5 * d.Amount;
    """,
    "Victim Count per Shelter": """
        SELECT
          s.Shelter_ID,
          s.Shelter_Name,
          COUNT(v.Victim_ID) AS Num_Victims
        FROM Shelter s
        LEFT JOIN Victim v ON s.Shelter_ID = v.Shelter_ID
        GROUP BY s.Shelter_ID, s.Shelter_Name;
    """,
    "Victims Over 30": """
        SELECT Victim_ID, Victim_Name, Age
        FROM Victim
        WHERE Age > 30
        ORDER BY Age DESC;
    """,
    "Severe Disasters over 55K": """
        SELECT
          Disaster_ID,
          Disaster_Name,
          Location,
          Severity,
          Start_Date,
          End_Date
        FROM Disaster
        WHERE Severity > 55000
        ORDER BY Severity DESC;
    """
}

# Export queries to Excel file
excel_file = "wildfire_results.xlsx"
with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
    for sheet_name, query in queries.items():
        cursor.execute(query)
        data = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
        df = pd.DataFrame(data, columns=columns)
        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

cursor.close()
conn.close()

# Auto-adjust column widths for better readability
wb = load_workbook(excel_file)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        # Set width with some padding
        ws.column_dimensions[col_letter].width = max_length + 2
wb.save(excel_file)

print(f"Queries saved to '{excel_file}' with formatted sheets!")
